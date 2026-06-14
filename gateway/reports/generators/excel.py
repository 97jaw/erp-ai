"""Excel report generator using openpyxl."""

from __future__ import annotations

import logging
import os
import uuid
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Brand colours (ARGB hex)
HEADER_BG = "FF1A1A2E"
HEADER_FG = "FFFFFFFF"
ACCENT_BG = "FF4A90D9"
ALT_ROW_BG = "FFF5F7FA"
BORDER_COLOR = "FFDEE2E6"

CURRENCY_FMT = '#,##0.00'


def _reports_dir() -> str:
    base = os.path.join(
        os.path.dirname(__file__), "..", "..", "..", "static", "reports"
    )
    path = os.path.abspath(base)
    os.makedirs(path, exist_ok=True)
    return path


def _report_title(report_type: str) -> str:
    titles = {
        "pandl": "Profit & Loss Statement",
        "balance_sheet": "Balance Sheet",
        "cash_flow": "Cash Flow Statement",
        "trial_balance": "Trial Balance",
        "project_expense": "Project Expense Summary",
    }
    return titles.get(report_type, report_type.replace("_", " ").title())


def _header_style() -> tuple[Font, PatternFill, Alignment]:
    font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    fill = PatternFill("solid", fgColor=HEADER_BG)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align


def _accent_style() -> tuple[Font, PatternFill, Alignment]:
    font = Font(name="Calibri", bold=True, color=HEADER_FG, size=10)
    fill = PatternFill("solid", fgColor=ACCENT_BG)
    align = Alignment(horizontal="left", vertical="center")
    return font, fill, align


def _thin_border() -> Border:
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value or ""))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _extract_summary(data: dict[str, Any]) -> dict[str, Any]:
    summary = {}
    for key in (
        "total_income", "total_expense", "net_profit", "net_loss",
        "total_debit", "total_credit", "balance",
        "total_expenses", "wo_amount", "spend_percentage",
    ):
        if key in data:
            summary[key] = data[key]
    for k, v in data.items():
        if isinstance(v, (int, float)) and k not in summary:
            summary[k] = v
    return summary


def _extract_lines(data: dict[str, Any]) -> list[dict[str, Any]]:
    for key in ("rows", "report_lines", "lines", "line_ids", "accounts", "items", "data"):
        val = data.get(key)
        if isinstance(val, list) and val:
            return val
    sections = data.get("sections") or data.get("groups") or []
    if isinstance(sections, list):
        lines = []
        for section in sections:
            if isinstance(section, dict):
                sub = section.get("lines") or section.get("children") or []
                lines.extend(sub if isinstance(sub, list) else [])
        if lines:
            return lines
    return []


class ExcelGenerator:
    """Generate a styled Excel report using openpyxl."""

    def generate(
        self,
        report_type: str,
        data: dict[str, Any],
        params: dict[str, Any],
    ) -> tuple[str, str]:
        """
        Build the Excel workbook.

        Returns (report_id, filepath).
        """
        report_id = uuid.uuid4().hex
        filename = f"{report_id}.xlsx"
        filepath = os.path.join(_reports_dir(), filename)

        wb = Workbook()

        # ------------------------------------------------------------------ #
        # Sheet 1: Summary
        # ------------------------------------------------------------------ #
        ws_sum = wb.active
        ws_sum.title = "Summary"

        title = _report_title(report_type)
        date_from = params.get("date_from", "")
        date_to = params.get("date_to", "")
        period = f"{date_from} to {date_to}" if date_from and date_to else "All periods"

        # Title row
        ws_sum.merge_cells("A1:B1")
        title_cell = ws_sum["A1"]
        title_cell.value = title
        title_cell.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=14)
        title_cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 30

        # Meta rows
        ws_sum["A2"] = "Period"
        ws_sum["B2"] = period
        ws_sum["A3"] = "Generated"
        ws_sum["B3"] = date.today().isoformat()

        for row in (2, 3):
            ws_sum.cell(row=row, column=1).font = Font(bold=True, size=10)
            ws_sum.cell(row=row, column=2).font = Font(size=10)

        ws_sum.append([])

        # Column headers
        hdr_row = ws_sum.max_row + 1
        ws_sum.cell(row=hdr_row, column=1).value = "Metric"
        ws_sum.cell(row=hdr_row, column=2).value = "Value"
        hfont, hfill, halign = _accent_style()
        for col in (1, 2):
            cell = ws_sum.cell(row=hdr_row, column=col)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign
            cell.border = _thin_border()

        # Summary data
        summary = _extract_summary(data)
        alt = False
        for key, val in summary.items():
            data_row = ws_sum.max_row + 1
            label_cell = ws_sum.cell(row=data_row, column=1, value=key.replace("_", " ").title())
            val_cell = ws_sum.cell(row=data_row, column=2, value=val)
            if isinstance(val, (int, float)):
                val_cell.number_format = CURRENCY_FMT
                val_cell.alignment = Alignment(horizontal="right")
            if alt:
                for col in (1, 2):
                    ws_sum.cell(row=data_row, column=col).fill = PatternFill(
                        "solid", fgColor=ALT_ROW_BG
                    )
            for col in (1, 2):
                ws_sum.cell(row=data_row, column=col).border = _thin_border()
                ws_sum.cell(row=data_row, column=col).font = Font(size=10)
            alt = not alt

        if not summary:
            ws_sum.append(["No summary data available."])

        _auto_width(ws_sum)

        # ------------------------------------------------------------------ #
        # Sheet 2: Detail
        # ------------------------------------------------------------------ #
        ws_det = wb.create_sheet("Detail")
        lines = _extract_lines(data)

        if lines and isinstance(lines[0], dict):
            cols = list(lines[0].keys())[:10]
            col_headers = [c.replace("_", " ").title() for c in cols]

            # Header row
            for ci, hdr in enumerate(col_headers, start=1):
                cell = ws_det.cell(row=1, column=ci, value=hdr)
                cell.font = Font(name="Calibri", bold=True, color=HEADER_FG, size=10)
                cell.fill = PatternFill("solid", fgColor=HEADER_BG)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _thin_border()
            ws_det.row_dimensions[1].height = 20

            # Data rows
            alt = False
            for line in lines[:500]:
                row_vals = []
                for col in cols:
                    row_vals.append(line.get(col))
                row_idx = ws_det.max_row + 1
                for ci, val in enumerate(row_vals, start=1):
                    cell = ws_det.cell(row=row_idx, column=ci, value=val)
                    cell.font = Font(size=9)
                    cell.border = _thin_border()
                    if isinstance(val, (int, float)):
                        cell.number_format = CURRENCY_FMT
                        cell.alignment = Alignment(horizontal="right")
                    if alt:
                        cell.fill = PatternFill("solid", fgColor=ALT_ROW_BG)
                alt = not alt

            _auto_width(ws_det)
            # Freeze header row
            ws_det.freeze_panes = "A2"
        else:
            ws_det["A1"] = "No detail data available."

        wb.save(filepath)
        logger.info("[ExcelGenerator] Generated %s → %s", report_type, filepath)
        return report_id, filepath
